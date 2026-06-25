"""CSV, Excel and text-report exporters."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from .logging_config import get_logger
from .statistics import Statistics

log = get_logger("exporters")

# Sheet name -> attribute on Statistics. The first seven are the required
# sheets (in the required order); the rest are useful extras.
_EXCEL_SHEETS: list[tuple[str, str]] = [
    ("Commits", "commits"),
    ("Pull Requests", "pull_requests"),
    ("Repositories", "repositories"),
    ("Organizations", "organizations"),
    ("Summary", "summary"),
    ("Yearly Stats", "commits_per_year"),
    ("Monthly Stats", "commits_per_month"),
    ("Per User", "per_user"),
    ("Commits Per Repo", "commits_per_repo"),
    ("By Organization", "commits_by_org"),
    ("By Email", "commits_by_email"),
    ("Top Repositories", "top_repositories"),
    ("PR Summary", "pr_summary"),
]

# CSV file name -> attribute on Statistics.
_CSV_FILES: list[tuple[str, str]] = [
    ("commits.csv", "commits"),
    ("pull_requests.csv", "pull_requests"),
    ("repositories.csv", "repositories"),
    ("organizations.csv", "organizations"),
    ("contribution_summary.csv", "summary"),
]


def export_csvs(output_dir: Path, stats: Statistics) -> list[Path]:
    """Write the required CSV files. Returns the paths written."""
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for filename, attr in _CSV_FILES:
        df: pd.DataFrame = getattr(stats, attr)
        path = output_dir / filename
        df.to_csv(path, index=False, encoding="utf-8-sig")
        written.append(path)
        log.info("wrote %s (%d rows)", path.name, len(df))
    return written


def export_excel(output_dir: Path, stats: Statistics, filename: str = "github_contributions.xlsx") -> Path:
    """Write a formatted multi-sheet Excel workbook."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / filename
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, attr in _EXCEL_SHEETS:
            df: pd.DataFrame = getattr(stats, attr)
            # Excel sheet names are capped at 31 chars and have reserved chars.
            safe_name = _safe_sheet_name(sheet_name)
            df.to_excel(writer, sheet_name=safe_name, index=False)
        _format_workbook(writer)
    log.info("wrote %s (%d sheets)", path.name, len(_EXCEL_SHEETS))
    return path


def export_summary_report(output_dir: Path, stats: Statistics, filename: str = "summary_report.txt") -> Path:
    """Write a human-readable plain-text summary report."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / filename
    lines: list[str] = []
    lines.append("=" * 64)
    lines.append("GITHUB CONTRIBUTION SUMMARY REPORT")
    lines.append("=" * 64)
    lines.append("")
    for key, value in stats.summary_dict.items():
        label = key.replace("_", " ").title()
        lines.append(f"{label:<34}: {value}")
    lines.append("")

    if not stats.per_user.empty:
        lines.append("-" * 64)
        lines.append("PER-USER BREAKDOWN")
        lines.append("-" * 64)
        lines.append(stats.per_user.to_string(index=False))
        lines.append("")

    if not stats.top_repositories.empty:
        lines.append("-" * 64)
        lines.append("TOP REPOSITORIES BY COMMITS")
        lines.append("-" * 64)
        lines.append(stats.top_repositories.to_string(index=False))
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    log.info("wrote %s", path.name)
    return path


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(name: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in name)
    return cleaned[:31]


def _format_workbook(writer: "pd.ExcelWriter") -> None:
    """Apply header styling, freeze panes, autofilter and column widths."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:  # pragma: no cover - openpyxl missing
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(vertical="center", horizontal="left")

    for worksheet in writer.sheets.values():
        max_col = worksheet.max_column
        max_row = worksheet.max_row
        if max_col == 0 or max_row == 0:
            continue
        # Style the header row.
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        worksheet.freeze_panes = "A2"
        last_col_letter = get_column_letter(max_col)
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

        # Approximate auto column widths (sampling rows keeps it fast).
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            longest = len(str(worksheet.cell(row=1, column=col_idx).value or ""))
            sample_rows = min(max_row, 200)
            for row_idx in range(2, sample_rows + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    longest = max(longest, len(str(value)))
            worksheet.column_dimensions[letter].width = min(max(longest + 2, 10), 60)
