"""Offline tests - no network access or GitHub tokens required.

Run directly:

    python tests/test_offline.py

or with pytest:

    pytest -q
"""

from __future__ import annotations

import asyncio
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

# Make the package importable when run as a plain script.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from github_contrib.client import GitHubClient, GitHubError  # noqa: E402
from github_contrib.commits import _commit_from_payload  # noqa: E402
from github_contrib.discovery import _repo_from_payload  # noqa: E402
from github_contrib.models import (  # noqa: E402
    CollectedData,
    CommitRecord,
    OrgRecord,
    PullRequestRecord,
    RepoRecord,
    parse_github_datetime,
)
from github_contrib.organizations import aggregate_organizations  # noqa: E402
from github_contrib.pull_requests import _pr_from_payload  # noqa: E402
from github_contrib.statistics import compute_statistics  # noqa: E402
from github_contrib.insights import (  # noqa: E402
    compute_insights,
    extract_themes,
    humanize_branch,
)
from github_contrib.config import token_env_candidates  # noqa: E402


# ---------------------------------------------------------------------------
# Fake aiohttp session for exercising the async client offline.
# ---------------------------------------------------------------------------

class _FakeResponse:
    def __init__(self, status, json_data=None, headers=None, text_data=""):
        self.status = status
        self._json = json_data
        self.headers = headers or {}
        self._text = text_data

    async def json(self):
        return self._json

    async def text(self):
        return self._text

    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return False


class _FakeSession:
    def __init__(self, responses):
        self._responses = list(responses)
        self.calls = []

    def request(self, method, url, params=None, headers=None, timeout=None):
        self.calls.append((method, url, params))
        resp = self._responses.pop(0)
        return resp() if callable(resp) else resp


def _client(session) -> GitHubClient:
    return GitHubClient(
        token="x",
        session=session,
        semaphore=asyncio.Semaphore(4),
        login="tester",
        max_retries=3,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_parse_datetime():
    assert parse_github_datetime(None) is None
    assert parse_github_datetime("") is None
    dt = parse_github_datetime("2023-05-01T12:34:56Z")
    assert dt is not None and dt.year == 2023 and dt.tzinfo is not None
    dt2 = parse_github_datetime("2024-01-02T03:04:05+00:00")
    assert dt2 is not None and dt2.month == 1
    assert parse_github_datetime("not-a-date") is None
    print("ok  test_parse_datetime")


def test_link_parsing():
    header = (
        '<https://api.github.com/x?page=2>; rel="next", '
        '<https://api.github.com/x?page=9>; rel="last"'
    )
    assert GitHubClient._parse_next_link(header) == "https://api.github.com/x?page=2"
    assert GitHubClient._parse_next_link('<https://x>; rel="last"') is None
    assert GitHubClient._parse_next_link(None) is None
    print("ok  test_link_parsing")


def test_payload_mappers():
    repo = _repo_from_payload(
        {
            "full_name": "acme/widget",
            "name": "widget",
            "owner": {"login": "acme", "type": "Organization"},
            "private": True,
            "default_branch": "main",
            "html_url": "https://github.com/acme/widget",
            "pushed_at": "2023-01-01T00:00:00Z",
        },
        "tester",
    )
    assert repo.organization == "acme" and repo.is_private and "tester" in repo.discovered_via

    commit = _commit_from_payload(
        {
            "sha": "abc123",
            "html_url": "https://github.com/acme/widget/commit/abc123",
            "commit": {
                "message": "Fix bug\n\nDetails",
                "author": {"name": "Mani", "email": "mani@example.com", "date": "2023-03-04T05:06:07Z"},
                "committer": {"name": "Mani", "email": "mani@example.com", "date": "2023-03-04T05:06:07Z"},
            },
            "author": {"login": "manisharai01"},
        },
        repo,
        "main",
        "manisharai01",
    )
    assert commit.sha == "abc123"
    assert commit.message_first_line == "Fix bug"
    assert commit.author_login == "manisharai01"
    assert commit.organization == "acme"

    pr = _pr_from_payload(
        {
            "number": 7,
            "title": "Add feature",
            "user": {"login": "manisharai01"},
            "state": "closed",
            "merged_at": "2023-04-01T00:00:00Z",
            "created_at": "2023-03-30T00:00:00Z",
            "base": {"ref": "main"},
            "head": {"ref": "feature"},
            "html_url": "https://github.com/acme/widget/pull/7",
        },
        repo,
    )
    assert pr.merged is True and pr.effective_state == "merged" and pr.number == 7
    print("ok  test_payload_mappers")


def test_async_pagination():
    page1 = _FakeResponse(
        200,
        json_data=[{"sha": "1"}, {"sha": "2"}, {"sha": "3"}],
        headers={"Link": '<https://api.github.com/next>; rel="next"'},
    )
    page2 = _FakeResponse(200, json_data=[{"sha": "4"}, {"sha": "5"}], headers={})
    session = _FakeSession([page1, page2])
    client = _client(session)

    async def run():
        return [item async for item in client.paginate("/repos/x/y/commits")]

    items = asyncio.run(run())
    assert len(items) == 5, items
    assert len(session.calls) == 2
    print("ok  test_async_pagination")


def test_async_rate_limit_retry():
    rate_limited = _FakeResponse(
        403,
        headers={"X-RateLimit-Remaining": "0", "X-RateLimit-Reset": "0"},
        text_data="API rate limit exceeded",
    )
    ok = _FakeResponse(200, json_data=[{"sha": "x"}], headers={})
    session = _FakeSession([rate_limited, ok])
    client = _client(session)

    async def fake_sleep(*_a, **_k):
        return None

    async def run():
        with patch("asyncio.sleep", fake_sleep):
            return [item async for item in client.paginate("/repos/x/y/commits")]

    items = asyncio.run(run())
    assert len(items) == 1
    assert client.rate_limit_waits == 1
    print("ok  test_async_rate_limit_retry")


def test_secondary_rate_limit_403_is_retried():
    """A secondary-rate-limit 403 (no Retry-After, budget intact) must back off
    and retry, not be misread as a permission error."""
    sec = _FakeResponse(
        403,
        headers={"X-RateLimit-Remaining": "4999"},
        text_data="You have exceeded a secondary rate limit. Please wait a few minutes.",
    )
    ok = _FakeResponse(200, json_data=[{"sha": "x"}], headers={})
    session = _FakeSession([sec, ok])
    client = _client(session)

    async def fake_sleep(*_a, **_k):
        return None

    async def run():
        with patch("asyncio.sleep", fake_sleep):
            return [item async for item in client.paginate("/repos/x/y/commits")]

    items = asyncio.run(run())
    assert len(items) == 1
    assert client.rate_limit_waits == 1
    print("ok  test_secondary_rate_limit_403_is_retried")


def test_permission_403_returns_empty_without_raising():
    """A genuine permission 403 on the first page yields [] (no exception)."""
    forbidden = _FakeResponse(
        403,
        headers={"X-RateLimit-Remaining": "4999"},
        text_data="Resource not accessible by personal access token",
    )
    session = _FakeSession([forbidden])
    client = _client(session)

    async def run():
        return [item async for item in client.paginate("/repos/x/secret/commits")]

    assert asyncio.run(run()) == []
    print("ok  test_permission_403_returns_empty_without_raising")


def test_midstream_truncation_raises():
    """If a later page comes back empty, paginate must raise rather than
    silently return a truncated list."""
    page1 = _FakeResponse(
        200,
        json_data=[{"sha": "1"}],
        headers={"Link": '<https://api.github.com/next>; rel="next"'},
    )
    page2_gone = _FakeResponse(404, json_data=None, headers={})
    session = _FakeSession([page1, page2_gone])
    client = _client(session)

    async def run():
        out = []
        async for item in client.paginate("/repos/x/y/commits"):
            out.append(item)
        return out

    raised = False
    try:
        asyncio.run(run())
    except GitHubError:
        raised = True
    assert raised, "expected GitHubError on mid-stream truncation"
    print("ok  test_midstream_truncation_raises")


def test_empty_status_is_not_error():
    empty = _FakeResponse(409, json_data=None, headers={}, text_data="empty repo")
    session = _FakeSession([empty])
    client = _client(session)

    async def run():
        return [item async for item in client.paginate("/repos/x/empty/commits")]

    assert asyncio.run(run()) == []
    print("ok  test_empty_status_is_not_error")


def _synthetic_data() -> CollectedData:
    repos = [
        RepoRecord(full_name="acme/widget", name="widget", owner="acme", organization="acme",
                   is_private=True, default_branch="main"),
        RepoRecord(full_name="manisharai01/personal", name="personal", owner="manisharai01",
                   organization="", default_branch="main"),
    ]
    repos[0].discovered_via = {"manisharai01"}
    repos[1].discovered_via = {"manisharai01"}

    def commit(repo, sha, when, login, email):
        return CommitRecord(
            repository=repo.name, full_name=repo.full_name, owner=repo.owner,
            organization=repo.organization, sha=sha, message=f"msg {sha}",
            message_first_line=f"msg {sha}", author_login=login, author_name=login,
            author_email=email, committer_name=login, committer_email=email,
            authored_date=when, committed_date=when, branch="main",
            url=f"https://github.com/{repo.full_name}/commit/{sha}",
        )

    commits = [
        commit(repos[0], "a1", datetime(2022, 1, 5, tzinfo=timezone.utc), "manisharai01", "m1@e.com"),
        commit(repos[0], "a2", datetime(2022, 2, 9, tzinfo=timezone.utc), "manisharai01", "m1@e.com"),
        commit(repos[0], "a3", datetime(2023, 6, 9, tzinfo=timezone.utc), "manisharai21", "m2@e.com"),
        commit(repos[1], "b1", datetime(2023, 7, 1, tzinfo=timezone.utc), "manisharai01", "m1@e.com"),
    ]
    prs = [
        PullRequestRecord(repository="widget", full_name="acme/widget", organization="acme",
                          number=1, title="PR1", author_login="manisharai01", state="closed",
                          merged=True, created_at=datetime(2023, 6, 1, tzinfo=timezone.utc),
                          updated_at=None, closed_at=None,
                          merged_at=datetime(2023, 6, 2, tzinfo=timezone.utc),
                          base_branch="main", head_branch="f1", url="u1"),
        PullRequestRecord(repository="widget", full_name="acme/widget", organization="acme",
                          number=2, title="PR2", author_login="manisharai21", state="open",
                          merged=False, created_at=datetime(2023, 6, 3, tzinfo=timezone.utc),
                          updated_at=None, closed_at=None, merged_at=None,
                          base_branch="main", head_branch="f2", url="u2"),
    ]
    member_orgs = [OrgRecord(login="acme", name="Acme Inc", is_member=True)]
    orgs = aggregate_organizations(repos, commits, prs, member_orgs)
    return CollectedData(repos=repos, commits=commits, pull_requests=prs, organizations=orgs)


def test_statistics():
    data = _synthetic_data()
    stats = compute_statistics(data)
    assert stats.summary_dict["total_lifetime_commits"] == 4
    assert stats.summary_dict["merged_pull_requests"] == 1
    assert stats.summary_dict["open_pull_requests"] == 1
    assert stats.summary_dict["repositories_contributed_to"] == 2
    assert not stats.commits_per_year.empty
    assert int(stats.commits_per_year["commits"].sum()) == 4
    assert not stats.commits_per_month.empty
    assert int(stats.commits_by_email["commits"].sum()) == 4
    # acme org should have 3 commits and 2 PRs.
    acme = next(o for o in data.organizations if o.login == "acme")
    assert acme.commit_count == 3 and acme.pr_count == 2 and acme.merged_pr_count == 1
    print("ok  test_statistics")


def test_exports_and_charts():
    data = _synthetic_data()
    stats = compute_statistics(data)
    from github_contrib.exporters import export_csvs, export_excel, export_summary_report

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp)
        export_csvs(out, stats)
        for name in ["commits.csv", "pull_requests.csv", "repositories.csv",
                     "organizations.csv", "contribution_summary.csv"]:
            p = out / name
            assert p.exists() and p.stat().st_size > 0, name

        xlsx = export_excel(out, stats)
        assert xlsx.exists()
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        required = {"Commits", "Pull Requests", "Repositories", "Organizations",
                    "Summary", "Yearly Stats", "Monthly Stats"}
        assert required.issubset(set(wb.sheetnames)), wb.sheetnames
        wb.close()

        export_summary_report(out, stats)
        assert (out / "summary_report.txt").exists()

        from github_contrib.charts import generate_all_charts
        charts = generate_all_charts(stats, out / "charts")
        for c in charts:
            assert c.exists() and c.stat().st_size > 0, c
    print("ok  test_exports_and_charts")


def test_exports_empty_data():
    """Empty collection must still produce valid (header-only) files."""
    stats = compute_statistics(CollectedData())
    from github_contrib.exporters import export_csvs, export_excel
    from github_contrib.charts import generate_all_charts

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp)
        export_csvs(out, stats)
        export_excel(out, stats)
        generate_all_charts(stats, out / "charts")
        assert (out / "commits.csv").exists()
        assert (out / "github_contributions.xlsx").exists()
    print("ok  test_exports_empty_data")


def test_humanize_branch():
    assert humanize_branch("AddHorsePhoto") == "Add Horse Photo"
    assert humanize_branch("403issue") == "403 issue"
    assert humanize_branch("feature/email-verified") == "Email verified"
    assert humanize_branch("fe-fat-money-loan") == "Fe fat money loan"
    assert humanize_branch("") == ""
    print("ok  test_humanize_branch")


def test_extract_themes():
    msgs = [
        "feat: add login page",
        "fix: login redirect bug",
        "refactor login validation",
        "update readme",
    ]
    themes, types = extract_themes(msgs)
    assert "login" in themes  # most frequent meaningful word
    assert types.get("feat") == 1 and types.get("fix") == 1
    print("ok  test_extract_themes")


def test_token_env_candidates():
    mapping = {"manisharai01": "GITHUB_TOKEN_1"}
    assert token_env_candidates("manisharai01", mapping)[0] == "GITHUB_TOKEN_1"
    cands = token_env_candidates("Octo-Cat", mapping)
    assert "GITHUB_TOKEN_OCTO_CAT" in cands and cands[-1] == "GITHUB_TOKEN"
    print("ok  test_token_env_candidates")


def test_insights_and_reports():
    data = _synthetic_data()
    insights = compute_insights(data)
    assert insights.total_active_days >= 1
    assert insights.repo_work, "expected per-repo work summaries"
    # acme/widget had a merged PR titled 'PR1' -> should surface as a highlight.
    widget = next((w for w in insights.repo_work if w.full_name == "acme/widget"), None)
    assert widget is not None
    assert any("PR1" in h for h in widget.highlights)

    stats = compute_statistics(data)
    from github_contrib.htmlreport import export_reports

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp)
        export_reports(out, stats, insights, stats.summary_dict, out / "charts", include_charts=False)
        html = (out / "report.html").read_text(encoding="utf-8")
        md = (out / "report.md").read_text(encoding="utf-8")
        assert (out / "report.html").exists() and (out / "report.md").exists()
        assert "Work breakdown by repository" in html
        assert "acme/widget" in html and "acme/widget" in md
        assert "<!doctype html>" in html.lower()
    print("ok  test_insights_and_reports")


def _all_tests():
    return [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]


def main() -> int:
    failures = 0
    for test in _all_tests():
        try:
            test()
        except Exception as exc:  # noqa: BLE001
            failures += 1
            print(f"FAIL {test.__name__}: {exc!r}")
            import traceback
            traceback.print_exc()
    total = len(_all_tests())
    print(f"\n{total - failures}/{total} tests passed.")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
